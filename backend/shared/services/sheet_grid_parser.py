"""
Sheet grid parser/extractor.

This module converts heterogeneous spreadsheet sources into a single, normalized representation:
- grid: rectangular 2D array (rows x cols), starting at A1 (0,0)
- merged_cells: list of MergeRange (0-based, inclusive)

Design principles:
- Keep parsing (I/O + format-specific quirks) separate from structure analysis.
- Prefer JSON-safe primitives in grid cells (string/number/bool/None).
- Trim only trailing empty rows/cols (bottom/right) so coordinates stay stable.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
import re
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from shared.models.sheet_grid import SheetGrid
from shared.models.structure_analysis import MergeRange


@dataclass(frozen=True)
class SheetGridParseOptions:
    """Options shared across parsers."""

    trim_trailing_empty: bool = True
    max_rows: Optional[int] = None
    max_cols: Optional[int] = None

    # Excel-specific
    excel_data_only: bool = True
    excel_fallback_to_formula: bool = True


class SheetGridParser:
    """Parsers for Excel and Google Sheets into SheetGrid."""

    # Currency hints used for Excel number_format heuristics
    _CURRENCY_SYMBOLS = ("₩", "¥", "￥", "$", "€", "£", "₹", "₽", "฿", "₫", "₱", "₪")
    _CURRENCY_UNIT_WORDS = ("원", "원화", "元", "人民币", "人民幣", "円")

    @classmethod
    def from_google_sheets_values(
        cls,
        values: List[List[Any]],
        *,
        merged_cells: Optional[Sequence[MergeRange]] = None,
        sheet_name: Optional[str] = None,
        options: Optional[SheetGridParseOptions] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> SheetGrid:
        """
        Build SheetGrid from Google Sheets "values" matrix (A1-anchored).

        Notes:
        - Values API returns ragged rows; we normalize to rectangular.
        - When merges are provided (from Sheets metadata), downstream can forward-fill.
        """
        opts = options or SheetGridParseOptions()
        grid = cls._normalize_grid(values, max_rows=opts.max_rows, max_cols=opts.max_cols)
        warnings: List[str] = []

        if opts.trim_trailing_empty:
            grid, trim_meta = cls._trim_trailing_empty(grid)
            if trim_meta.get("trimmed"):
                warnings.append("Trailing empty rows/cols trimmed")

        rows = len(grid)
        cols = max((len(r) for r in grid), default=0)
        merges = cls._clip_merge_ranges(list(merged_cells or []), rows=rows, cols=cols)

        return SheetGrid(
            source="google_sheets",
            sheet_name=sheet_name,
            grid=grid,
            merged_cells=merges,
            metadata={"rows": rows, "cols": cols, **(metadata or {})},
            warnings=warnings,
        )

    @classmethod
    def merged_cells_from_google_metadata(
        cls,
        sheets_metadata: Dict[str, Any],
        *,
        worksheet_name: Optional[str] = None,
        sheet_id: Optional[int] = None,
    ) -> List[MergeRange]:
        """
        Extract merged ranges from Google Sheets "spreadsheets.get" metadata JSON.

        The API uses 0-based indexes with exclusive end bounds; we convert to inclusive MergeRange.
        """
        sheets = sheets_metadata.get("sheets", [])
        target = None
        if sheet_id is not None:
            for s in sheets:
                if int(s.get("properties", {}).get("sheetId", -1)) == int(sheet_id):
                    target = s
                    break
        if target is None and worksheet_name:
            for s in sheets:
                if str(s.get("properties", {}).get("title", "")) == worksheet_name:
                    target = s
                    break
        if target is None and sheets:
            target = sheets[0]

        merges = []
        for m in target.get("merges", []) if isinstance(target, dict) else []:
            try:
                sr = int(m.get("startRowIndex", 0))
                er = int(m.get("endRowIndex", 0))
                sc = int(m.get("startColumnIndex", 0))
                ec = int(m.get("endColumnIndex", 0))
            except (TypeError, ValueError):
                continue
            if er <= sr or ec <= sc:
                continue
            merges.append(MergeRange(top=sr, left=sc, bottom=er - 1, right=ec - 1))
        return merges

    @classmethod
    def from_excel_bytes(
        cls,
        xlsx_bytes: bytes,
        *,
        sheet_name: Optional[str] = None,
        options: Optional[SheetGridParseOptions] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> SheetGrid:
        """
        Parse an .xlsx file into SheetGrid.

        Requires optional dependency: openpyxl.
        """
        opts = options or SheetGridParseOptions()
        warnings: List[str] = []

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:  # pragma: no cover
            raise RuntimeError(
                "Excel parsing requires 'openpyxl'. Install it to enable .xlsx support."
            ) from e

        wb = load_workbook(
            filename=BytesIO(xlsx_bytes),
            data_only=bool(opts.excel_data_only),
            read_only=False,
        )
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        # Bounds
        max_row = int(ws.max_row or 0)
        max_col = int(ws.max_column or 0)
        if opts.max_rows is not None:
            max_row = min(max_row, int(opts.max_rows))
        if opts.max_cols is not None:
            max_col = min(max_col, int(opts.max_cols))

        # Merged ranges (1-based in openpyxl)
        merges: List[MergeRange] = []
        try:
            for cr in list(getattr(ws.merged_cells, "ranges", [])):
                mr = MergeRange(
                    top=int(cr.min_row) - 1,
                    left=int(cr.min_col) - 1,
                    bottom=int(cr.max_row) - 1,
                    right=int(cr.max_col) - 1,
                )
                merges.append(mr)
        except Exception:
            warnings.append("Failed to read merged cell ranges from workbook")

        # Read cells (keep A1 anchor; downstream structure analyzer expects leading blanks too)
        grid: List[List[Any]] = []
        if max_row <= 0 or max_col <= 0:
            return SheetGrid(
                source="excel",
                sheet_name=str(ws.title),
                grid=[],
                merged_cells=[],
                metadata={"rows": 0, "cols": 0, **(metadata or {})},
                warnings=warnings,
            )

        # Iterate row-major for cache locality
        for r in range(1, max_row + 1):
            row_out: List[Any] = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                row_out.append(
                    cls._excel_cell_to_display_value(cell, fallback_to_formula=opts.excel_fallback_to_formula)
                )
            grid.append(row_out)

        if opts.trim_trailing_empty:
            # Preserve trailing empty rows/cols that are part of merged ranges (e.g. A3:A4 where A4 is blank).
            trimmed_merges = cls._clip_merge_ranges(merges, rows=max_row, cols=max_col)
            min_rows = (max((m.bottom for m in trimmed_merges), default=-1) + 1) if trimmed_merges else 0
            min_cols = (max((m.right for m in trimmed_merges), default=-1) + 1) if trimmed_merges else 0
            grid, trim_meta = cls._trim_trailing_empty(grid, min_rows=min_rows, min_cols=min_cols)
            if trim_meta.get("trimmed"):
                warnings.append("Trailing empty rows/cols trimmed")

        rows = len(grid)
        cols = max((len(r) for r in grid), default=0)
        merges = cls._clip_merge_ranges(merges, rows=rows, cols=cols)

        return SheetGrid(
            source="excel",
            sheet_name=str(ws.title),
            grid=grid,
            merged_cells=merges,
            metadata={"rows": rows, "cols": cols, **(metadata or {})},
            warnings=warnings,
        )

    # -------------------------
    # Normalization helpers
    # -------------------------

    @classmethod
    def _normalize_grid(
        cls, grid: List[List[Any]], *, max_rows: Optional[int], max_cols: Optional[int]
    ) -> List[List[Any]]:
        if not grid:
            return []

        rows = grid[: max_rows] if max_rows is not None else grid
        width = max((len(r) for r in rows), default=0)
        if max_cols is not None:
            width = min(width, int(max_cols))

        out: List[List[Any]] = []
        for row in rows:
            sliced = row[:width]
            if len(sliced) < width:
                sliced = list(sliced) + [""] * (width - len(sliced))
            out.append([cls._json_safe_cell(v) for v in sliced])
        return out

    @staticmethod
    def _json_safe_cell(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, (str, int, float, bool)):
            return value
        if isinstance(value, Decimal):
            return str(value)
        if isinstance(value, (datetime, date, time)):
            # ISO 8601; downstream type inference supports common date/datetime patterns.
            if isinstance(value, datetime):
                return value.isoformat(sep=" ")
            return value.isoformat()
        return str(value)

    @classmethod
    def _trim_trailing_empty(
        cls,
        grid: List[List[Any]],
        *,
        min_rows: int = 0,
        min_cols: int = 0,
    ) -> Tuple[List[List[Any]], Dict[str, Any]]:
        if not grid:
            return grid, {"trimmed": False}

        rows = len(grid)
        cols = max((len(r) for r in grid), default=0)

        def is_blank(v: Any) -> bool:
            return v is None or str(v).strip() == ""

        bottom = rows - 1
        while bottom >= 0 and all(is_blank(v) for v in grid[bottom]):
            bottom -= 1
        bottom = max(bottom, int(min_rows) - 1)

        right = cols - 1
        while right >= 0:
            if all(is_blank(grid[r][right]) for r in range(0, bottom + 1)):
                right -= 1
            else:
                break
        right = max(right, int(min_cols) - 1)

        trimmed = (bottom != rows - 1) or (right != cols - 1)
        new_grid = [row[: right + 1] for row in grid[: bottom + 1]]
        return new_grid, {"trimmed": trimmed, "rows": bottom + 1, "cols": right + 1}

    @classmethod
    def _clip_merge_ranges(cls, merges: List[MergeRange], *, rows: int, cols: int) -> List[MergeRange]:
        if not merges or rows <= 0 or cols <= 0:
            return []
        out: List[MergeRange] = []
        for m in merges:
            top = max(0, min(rows - 1, m.top))
            left = max(0, min(cols - 1, m.left))
            bottom = max(0, min(rows - 1, m.bottom))
            right = max(0, min(cols - 1, m.right))
            if bottom < top or right < left:
                continue
            # Ignore "merges" that are effectively single cells
            if top == bottom and left == right:
                continue
            out.append(MergeRange(top=top, left=left, bottom=bottom, right=right))
        return out

    # -------------------------
    # Excel display formatting
    # -------------------------

    @classmethod
    def _excel_cell_to_display_value(cls, cell: Any, *, fallback_to_formula: bool) -> Any:
        """
        Best-effort conversion of an openpyxl cell into a display-like value.

        Why: currency/date formats live in number_format; raw numeric values lose the symbol.
        """
        value = getattr(cell, "value", None)
        if value is None:
            # If computed values are missing, keep formula string as a last resort
            if fallback_to_formula and getattr(cell, "data_type", None) == "f":
                f = getattr(cell, "value", None)
                if isinstance(f, str) and f.startswith("="):
                    return f
            return ""

        if isinstance(value, bool):
            return "true" if value else "false"

        if isinstance(value, datetime):
            # Excel often stores date as datetime midnight
            if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
                return value.date().isoformat()
            return value.isoformat(sep=" ")

        if isinstance(value, date):
            return value.isoformat()

        if isinstance(value, time):
            return value.isoformat()

        if isinstance(value, (int, float, Decimal)):
            fmt = str(getattr(cell, "number_format", "") or "")
            return cls._format_excel_number(value, fmt)

        return cls._json_safe_cell(value)

    @classmethod
    def _format_excel_number(cls, value: Any, fmt: str) -> str:
        s = str(fmt or "")

        # Percent formats: store as fraction in Excel (0.3) -> display 30%
        if "%" in s:
            try:
                return f"{float(value) * 100:.0f}%"
            except Exception:
                return f"{value}%"

        currency_prefix, currency_suffix = cls._detect_currency_affixes(s)
        decimal_places = cls._infer_decimal_places_from_format(s)
        use_thousands = "," in s

        try:
            num = float(value)
        except Exception:
            return str(value)

        # Basic numeric formatting (best-effort)
        if decimal_places is None:
            # Preserve integers cleanly, otherwise keep up to 6 decimals without trailing zeros
            if abs(num - int(num)) < 1e-9:
                core = f"{int(num):,}" if use_thousands else str(int(num))
            else:
                core = f"{num:,.6f}" if use_thousands else f"{num:.6f}"
                core = core.rstrip("0").rstrip(".")
        else:
            if decimal_places <= 0:
                core = f"{num:,.0f}" if use_thousands else f"{num:.0f}"
            else:
                core = f"{num:,.{decimal_places}f}" if use_thousands else f"{num:.{decimal_places}f}"

        return f"{currency_prefix}{core}{currency_suffix}"

    @classmethod
    def _detect_currency_affixes(cls, fmt: str) -> Tuple[str, str]:
        s = fmt

        # Prefer explicit unit words embedded in the format (Korean/Chinese/Japanese)
        for word in cls._CURRENCY_UNIT_WORDS:
            if word in s:
                return "", word

        # Common symbols (prefix)
        for sym in cls._CURRENCY_SYMBOLS:
            if sym in s:
                return sym, ""

        # Bracket notation: [$¥-ja-JP], [$￥-804], etc.
        m = re.search(r"\[\$([^\]-]+)", s)
        if m:
            token = m.group(1).strip()
            if token in cls._CURRENCY_SYMBOLS:
                return token, ""
            # Sometimes token is a code like USD/EUR/RMB; keep as suffix for validator compatibility
            if re.fullmatch(r"[A-Za-z]{3}", token):
                return "", f" {token.upper()}"

        return "", ""

    @staticmethod
    def _infer_decimal_places_from_format(fmt: str) -> Optional[int]:
        # Take the positive section before ';'
        section = fmt.split(";", 1)[0]

        # Remove quoted strings and bracket directives (locale/currency)
        section = re.sub(r"\"[^\"]*\"", "", section)
        section = re.sub(r"\[[^\]]+\]", "", section)

        # Find the first numeric pattern containing a decimal point
        m = re.search(r"[#0]+\.([0#]+)", section)
        if not m:
            # If there's no explicit decimal point, assume integer format.
            return 0 if re.search(r"[#0]", section) else None

        decimals = m.group(1)
        # Count mandatory decimals (0); optional (#) still means "up to"
        mandatory = decimals.count("0")
        optional = decimals.count("#")
        if mandatory > 0:
            return mandatory
        if optional > 0:
            return len(decimals)
        return 0
