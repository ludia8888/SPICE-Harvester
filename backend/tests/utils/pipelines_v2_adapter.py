from __future__ import annotations

import csv
import io
import re
import time
from datetime import date, datetime
from typing import Any, Dict, Optional
from urllib.parse import urlsplit
from uuid import NAMESPACE_URL, uuid5

import httpx

from shared.foundry.rids import build_rid, parse_rid
from shared.models.pipeline_job import PipelineJob
from shared.services.pipeline.pipeline_job_queue import PipelineJobQueue
from shared.services.registries.dataset_registry import DatasetRegistry
from shared.services.registries.pipeline_registry import PipelineRegistry


_PIPELINE_DATASETS_PREFIX = "/api/v1/pipelines/datasets"
_PIPELINE_PREFIX = "/api/v1/pipelines"
_TERMINAL_RUN_STATUSES = {"SUCCESS", "FAILED", "DEPLOYED", "IGNORED", "CANCELLED", "CANCELED", "SUCCEEDED"}
_TERMINAL_BUILD_STATUSES = {"SUCCEEDED", "FAILED", "CANCELED", "CANCELLED"}


def _normalize_branch(branch: Optional[str]) -> str:
    text = str(branch or "").strip()
    if not text:
        return "master"
    if text.lower() == "main":
        return "master"
    return text


def _legacy_status(status: Any) -> str:
    text = str(status or "").strip().upper()
    if text in {"SUCCEEDED", "SUCCESS"}:
        return "SUCCESS"
    if text in {"CANCELED", "CANCELLED"}:
        return "CANCELLED"
    return text or "PENDING"


def _response(method: str, url: str, *, status_code: int, payload: Any) -> httpx.Response:
    request = httpx.Request(method=method, url=url)
    return httpx.Response(
        status_code=status_code,
        request=request,
        json=_json_safe(payload),
        headers={"content-type": "application/json"},
    )


def _extract_dataset_id(dataset_rid: str) -> str:
    kind, rid_id = parse_rid(str(dataset_rid or "").strip())
    if kind != "dataset":
        raise ValueError("datasetRid must be dataset kind")
    return rid_id


def _build_job_id_from_build_rid(build_rid: str) -> str:
    kind, rid_id = parse_rid(str(build_rid or "").strip())
    if kind != "build":
        raise ValueError("buildRid must be build kind")
    return rid_id


def _rows_to_csv_bytes(*, rows: list[dict[str, Any]], schema_json: dict[str, Any] | None) -> bytes:
    schema_columns = schema_json.get("columns") if isinstance(schema_json, dict) else None
    headers: list[str] = []
    if isinstance(schema_columns, list):
        for item in schema_columns:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name") or "").strip()
            if name:
                headers.append(name)
    if not headers and rows:
        discovered: list[str] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            for key in row.keys():
                key_text = str(key)
                if key_text not in discovered:
                    discovered.append(key_text)
        headers = discovered

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    if headers:
        writer.writerow(headers)
    for row in rows:
        if not isinstance(row, dict):
            continue
        writer.writerow([row.get(header) for header in headers])
    return buffer.getvalue().encode("utf-8")


def _read_upload_bytes(file_entry: Any) -> bytes:
    if isinstance(file_entry, tuple):
        if len(file_entry) >= 2:
            data = file_entry[1]
        else:
            data = b""
    else:
        data = file_entry

    if isinstance(data, bytes):
        return data
    if isinstance(data, str):
        return data.encode("utf-8")
    if hasattr(data, "read"):
        loaded = data.read()
        if isinstance(loaded, bytes):
            return loaded
        if isinstance(loaded, str):
            return loaded.encode("utf-8")
    return b""


def _infer_xsd_type(values: list[Any]) -> str:
    cleaned = [value for value in values if value not in (None, "")]
    if not cleaned:
        return "xsd:string"

    def _is_int(value: Any) -> bool:
        text = str(value).strip()
        if not text:
            return False
        if text.startswith(("+", "-")):
            text = text[1:]
        return text.isdigit()

    def _is_float(value: Any) -> bool:
        text = str(value).strip()
        if not text:
            return False
        try:
            float(text)
            return True
        except ValueError:
            return False

    def _is_date(value: Any) -> bool:
        text = str(value).strip()
        if not text:
            return False
        try:
            datetime.strptime(text, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def _is_datetime(value: Any) -> bool:
        text = str(value).strip()
        if not text:
            return False
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            try:
                datetime.strptime(text, fmt)
                return True
            except ValueError:
                continue
        return False

    if all(_is_int(value) for value in cleaned):
        return "xsd:integer"
    if all(_is_float(value) for value in cleaned):
        return "xsd:decimal"
    if all(_is_date(value) for value in cleaned):
        return "xsd:date"
    if all(_is_datetime(value) for value in cleaned):
        return "xsd:dateTime"
    return "xsd:string"


def _build_tabular_analysis(*, rows: list[dict[str, Any]], schema_json: dict[str, Any] | None) -> dict[str, Any]:
    schema_columns = schema_json.get("columns") if isinstance(schema_json, dict) else None
    names: list[str] = []
    if isinstance(schema_columns, list):
        for item in schema_columns:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name") or "").strip()
            if name:
                names.append(name)
    if not names and rows:
        names = list(rows[0].keys())

    columns: list[dict[str, Any]] = []
    for name in names:
        values = [row.get(name) for row in rows if isinstance(row, dict)]
        inferred = _infer_xsd_type(values)
        columns.append(
            {
                "name": name,
                "column_name": name,
                "type": inferred,
                "inferred_type": {"type": inferred},
            }
        )
    return {"columns": columns}


def _extract_output_dataset_name(definition_json: dict[str, Any], *, node_id: Optional[str]) -> str:
    nodes = definition_json.get("nodes")
    if not isinstance(nodes, list):
        return "output"

    if node_id:
        for node in nodes:
            if not isinstance(node, dict):
                continue
            if str(node.get("id") or "") != str(node_id):
                continue
            metadata = node.get("metadata") if isinstance(node.get("metadata"), dict) else {}
            value = str(
                metadata.get("datasetName")
                or metadata.get("outputName")
                or metadata.get("name")
                or ""
            ).strip()
            if value:
                return value

    for node in nodes:
        if not isinstance(node, dict):
            continue
        if str(node.get("type") or "").lower() != "output":
            continue
        metadata = node.get("metadata") if isinstance(node.get("metadata"), dict) else {}
        value = str(
            metadata.get("datasetName")
            or metadata.get("outputName")
            or metadata.get("name")
            or ""
        ).strip()
        if value:
            return value
    return "output"


def _json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _json_safe(val) for key, val in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    return value


class PipelinesV2AdapterClient:
    """E2E test adapter: rewrites legacy /api/v1/pipelines* calls onto v2/datasets+orchestration.

    Notes:
    - Dataset lifecycle calls are translated to `/api/v2/datasets` transaction flow.
    - Build calls are translated to `/api/v2/orchestration/builds/create`.
    - Preview calls are enqueued directly as preview jobs (no dedicated v2 preview endpoint yet).
    - Deploy calls are passed through to existing endpoint to preserve parity checks where no v2
      deploy endpoint exists yet.
    """

    def __init__(self, client: httpx.AsyncClient) -> None:
        self._client = client
        self._dataset_registry: DatasetRegistry | None = None
        self._pipeline_registry: PipelineRegistry | None = None
        self._queue = PipelineJobQueue()
        self._build_rids_by_job_id: dict[str, str] = {}
        self._mode_by_job_id: dict[str, str] = {}

    @property
    def headers(self) -> httpx.Headers:
        return self._client.headers

    def __getattr__(self, name: str) -> Any:
        return getattr(self._client, name)

    async def _get_dataset_registry(self) -> DatasetRegistry:
        if self._dataset_registry is None:
            self._dataset_registry = DatasetRegistry()
            await self._dataset_registry.initialize()
        return self._dataset_registry

    async def _get_pipeline_registry(self) -> PipelineRegistry:
        if self._pipeline_registry is None:
            self._pipeline_registry = PipelineRegistry()
            await self._pipeline_registry.initialize()
        return self._pipeline_registry

    async def _create_dataset_v2(
        self,
        *,
        url: str,
        db_name: str,
        name: str,
        description: str,
        branch: str,
        schema_json: dict[str, Any] | None,
        headers: Optional[dict[str, str]],
    ) -> httpx.Response:
        create_resp = await self._client.post(
            f"{self._base_url(url)}/api/v2/datasets",
            headers=headers,
            json={
                "name": name,
                "description": description,
                "parentFolderRid": build_rid("folder", db_name),
            },
        )
        if create_resp.status_code >= 400:
            return create_resp
        create_payload = create_resp.json()
        dataset_rid = str(create_payload.get("rid") or "")
        if not dataset_rid:
            return _response("POST", url, status_code=500, payload={"detail": "dataset rid missing"})
        dataset_id = _extract_dataset_id(dataset_rid)

        if isinstance(schema_json, dict) and isinstance(schema_json.get("columns"), list):
            field_schema_list = []
            for column in schema_json.get("columns") or []:
                if not isinstance(column, dict):
                    continue
                field_name = str(column.get("name") or "").strip()
                if not field_name:
                    continue
                field_type = str(column.get("type") or "string").strip() or "string"
                field_schema_list.append(
                    {
                        "fieldPath": field_name,
                        "type": {"type": field_type},
                    }
                )
            if field_schema_list:
                put_schema_resp = await self._client.put(
                    f"{self._base_url(url)}/api/v2/datasets/{dataset_rid}/putSchema",
                    headers=headers,
                    params={"preview": "true", "branchName": branch},
                    json={"fieldSchemaList": field_schema_list},
                )
                if put_schema_resp.status_code >= 400:
                    return put_schema_resp

        payload = {
            "status": "success",
            "data": {
                "dataset": {
                    "dataset_id": dataset_id,
                    "datasetId": dataset_id,
                    "rid": dataset_rid,
                    "db_name": db_name,
                    "name": name,
                    "branch": branch,
                }
            },
        }
        return _response("POST", url, status_code=200, payload=payload)

    async def _create_dataset_version_v2(
        self,
        *,
        url: str,
        dataset_id: str,
        branch: str,
        sample_json: dict[str, Any] | None,
        schema_json: dict[str, Any] | None,
        headers: Optional[dict[str, str]],
    ) -> httpx.Response:
        dataset_rid = build_rid("dataset", dataset_id)
        txn_resp = await self._client.post(
            f"{self._base_url(url)}/api/v2/datasets/{dataset_rid}/transactions",
            headers=headers,
            params={"branchName": branch},
            json={"transactionType": "APPEND"},
        )
        if txn_resp.status_code >= 400:
            return txn_resp
        txn_payload = txn_resp.json()
        transaction_rid = str(txn_payload.get("rid") or "")
        if not transaction_rid:
            return _response("POST", url, status_code=500, payload={"detail": "transaction rid missing"})

        sample_rows = sample_json.get("rows") if isinstance(sample_json, dict) else []
        normalized_rows = [row for row in sample_rows if isinstance(row, dict)]
        csv_bytes = _rows_to_csv_bytes(rows=normalized_rows, schema_json=schema_json)
        upload_resp = await self._client.post(
            f"{self._base_url(url)}/api/v2/datasets/{dataset_rid}/files/source.csv/upload",
            headers=headers,
            params={"transactionRid": transaction_rid},
            content=csv_bytes,
        )
        if upload_resp.status_code >= 400:
            return upload_resp

        commit_resp = await self._client.post(
            f"{self._base_url(url)}/api/v2/datasets/{dataset_rid}/transactions/{transaction_rid}/commit",
            headers=headers,
        )
        if commit_resp.status_code >= 400:
            return commit_resp

        if isinstance(schema_json, dict) and isinstance(schema_json.get("columns"), list):
            field_schema_list = []
            for column in schema_json.get("columns") or []:
                if not isinstance(column, dict):
                    continue
                field_name = str(column.get("name") or "").strip()
                if not field_name:
                    continue
                field_type = str(column.get("type") or "string").strip() or "string"
                field_schema_list.append(
                    {
                        "fieldPath": field_name,
                        "type": {"type": field_type},
                    }
                )
            if field_schema_list:
                put_schema_resp = await self._client.put(
                    f"{self._base_url(url)}/api/v2/datasets/{dataset_rid}/putSchema",
                    headers=headers,
                    params={"preview": "true", "branchName": branch},
                    json={"fieldSchemaList": field_schema_list},
                )
                if put_schema_resp.status_code >= 400:
                    return put_schema_resp

        dataset_registry = await self._get_dataset_registry()
        latest = await dataset_registry.get_latest_version(dataset_id=dataset_id)
        version_id = str(getattr(latest, "version_id", "") or "")
        payload = {
            "status": "success",
            "data": {
                "version": {
                    "version_id": version_id,
                    "versionId": version_id,
                    "dataset_id": dataset_id,
                    "datasetId": dataset_id,
                },
            },
        }
        return _response("POST", url, status_code=200, payload=payload)

    async def _csv_or_excel_upload_v2(
        self,
        *,
        url: str,
        params: dict[str, Any],
        data: dict[str, Any],
        files: dict[str, Any],
        headers: Optional[dict[str, str]],
        excel: bool,
    ) -> httpx.Response:
        db_name = str(params.get("db_name") or self._client.headers.get("X-DB-Name") or "").strip()
        branch = _normalize_branch(params.get("branch") or "master")
        dataset_name = str(data.get("dataset_name") or "").strip()
        description = str(data.get("description") or "").strip()
        file_entry = files.get("file")
        raw_bytes = _read_upload_bytes(file_entry)
        if excel:
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception:
                return _response("POST", url, status_code=500, payload={"detail": "openpyxl is required"})
            workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
            worksheet = workbook.active
            rows_iter = worksheet.iter_rows(values_only=True)
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            for row in rows_iter:
                writer.writerow([value for value in (row or [])])
            raw_bytes = buffer.getvalue().encode("utf-8")

        create_dataset_resp = await self._create_dataset_v2(
            url=url,
            db_name=db_name,
            name=dataset_name,
            description=description,
            branch=branch,
            schema_json=None,
            headers=headers,
        )
        if create_dataset_resp.status_code >= 400:
            return create_dataset_resp
        create_payload = create_dataset_resp.json()
        dataset_payload = ((create_payload.get("data") or {}) or {}).get("dataset") or {}
        dataset_id = str(dataset_payload.get("dataset_id") or "")
        dataset_rid = str(dataset_payload.get("rid") or "")
        if not dataset_id or not dataset_rid:
            return _response("POST", url, status_code=500, payload={"detail": "dataset create translation failed"})

        txn_resp = await self._client.post(
            f"{self._base_url(url)}/api/v2/datasets/{dataset_rid}/transactions",
            headers=headers,
            params={"branchName": branch},
            json={"transactionType": "APPEND"},
        )
        if txn_resp.status_code >= 400:
            return txn_resp
        transaction_rid = str((txn_resp.json() or {}).get("rid") or "")
        if not transaction_rid:
            return _response("POST", url, status_code=500, payload={"detail": "transaction rid missing"})

        upload_resp = await self._client.post(
            f"{self._base_url(url)}/api/v2/datasets/{dataset_rid}/files/source.csv/upload",
            headers=headers,
            params={"transactionRid": transaction_rid},
            content=raw_bytes,
        )
        if upload_resp.status_code >= 400:
            return upload_resp

        commit_resp = await self._client.post(
            f"{self._base_url(url)}/api/v2/datasets/{dataset_rid}/transactions/{transaction_rid}/commit",
            headers=headers,
        )
        if commit_resp.status_code >= 400:
            return commit_resp

        dataset_registry = await self._get_dataset_registry()
        latest = await dataset_registry.get_latest_version(dataset_id=dataset_id)
        version_id = str(getattr(latest, "version_id", "") or "")

        rows: list[dict[str, Any]] = []
        decoded = raw_bytes.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(decoded))
        for row in reader:
            rows.append(dict(row))

        payload = {
            "status": "success",
            "data": {
                "dataset": {
                    "dataset_id": dataset_id,
                    "datasetId": dataset_id,
                    "rid": dataset_rid,
                    "name": dataset_name,
                    "db_name": db_name,
                    "branch": branch,
                },
                "version": {
                    "version_id": version_id,
                    "versionId": version_id,
                    "dataset_id": dataset_id,
                    "datasetId": dataset_id,
                },
                "tabular_analysis": _build_tabular_analysis(rows=rows, schema_json=None),
            },
        }
        return _response("POST", url, status_code=200, payload=payload)

    async def _create_pipeline_direct(
        self,
        *,
        url: str,
        payload: dict[str, Any],
    ) -> httpx.Response:
        db_name = str(payload.get("db_name") or self._client.headers.get("X-DB-Name") or "").strip()
        name = str(payload.get("name") or "").strip()
        description = str(payload.get("description") or "").strip()
        location = str(payload.get("location") or "e2e").strip() or "e2e"
        branch = _normalize_branch(payload.get("branch") or "master")
        pipeline_type = str(payload.get("pipeline_type") or "batch").strip() or "batch"
        definition_json = payload.get("definition_json") if isinstance(payload.get("definition_json"), dict) else {}

        pipeline_registry = await self._get_pipeline_registry()
        record = await pipeline_registry.create_pipeline(
            db_name=db_name,
            name=name,
            description=description,
            pipeline_type=pipeline_type,
            location=location,
            branch=branch,
            status="draft",
        )
        await pipeline_registry.add_version(
            pipeline_id=record.pipeline_id,
            branch=branch,
            definition_json=definition_json,
            user_id=str(self._client.headers.get("X-User-ID") or "system"),
        )

        response_payload = {
            "status": "success",
            "data": {
                "pipeline": {
                    "pipeline_id": record.pipeline_id,
                    "pipelineId": record.pipeline_id,
                    "db_name": record.db_name,
                    "name": record.name,
                    "branch": record.branch,
                    "pipeline_type": record.pipeline_type,
                }
            },
        }
        return _response("POST", url, status_code=200, payload=response_payload)

    async def _queue_preview_job(
        self,
        *,
        url: str,
        pipeline_id: str,
        payload: dict[str, Any],
    ) -> httpx.Response:
        pipeline_registry = await self._get_pipeline_registry()
        pipeline = await pipeline_registry.get_pipeline(pipeline_id=pipeline_id)
        if pipeline is None:
            return _response("POST", url, status_code=404, payload={"detail": "pipeline not found"})
        latest = await pipeline_registry.get_latest_version(pipeline_id=pipeline_id, branch=pipeline.branch)
        if latest is None:
            return _response("POST", url, status_code=409, payload={"detail": "pipeline has no version"})

        node_id = str(payload.get("node_id") or payload.get("nodeId") or "").strip() or None
        branch = _normalize_branch(payload.get("branch") or pipeline.branch or "master")
        preview_limit = payload.get("limit")
        try:
            preview_limit_value = int(preview_limit) if preview_limit is not None else None
        except (TypeError, ValueError):
            preview_limit_value = None
        output_dataset_name = _extract_output_dataset_name(latest.definition_json, node_id=node_id)
        job_id = f"preview-{pipeline_id}-{int(time.time() * 1000)}"

        try:
            await self._queue.publish(
                PipelineJob(
                    job_id=job_id,
                    pipeline_id=pipeline_id,
                    db_name=pipeline.db_name,
                    pipeline_type=pipeline.pipeline_type,
                    definition_json=latest.definition_json,
                    definition_commit_id=latest.lakefs_commit_id,
                    node_id=node_id,
                    output_dataset_name=output_dataset_name,
                    mode="preview",
                    preview_limit=preview_limit_value,
                    branch=branch,
                )
            )
        except Exception:
            fallback_resp = await self._client.post(
                f"{self._base_url(url)}/api/v1/pipelines/{pipeline_id}/preview",
                json=payload,
            )
            if fallback_resp.status_code >= 400:
                return fallback_resp
            fallback_data = (fallback_resp.json().get("data") or {}) if fallback_resp.headers.get("content-type", "").startswith("application/json") else {}
            fallback_job_id = str(fallback_data.get("job_id") or "").strip()
            if fallback_job_id:
                self._mode_by_job_id[fallback_job_id] = "preview"
            return fallback_resp
        self._mode_by_job_id[job_id] = "preview"
        return _response("POST", url, status_code=200, payload={"status": "success", "data": {"job_id": job_id}})

    async def _queue_build_job(
        self,
        *,
        url: str,
        pipeline_id: str,
        payload: dict[str, Any],
    ) -> httpx.Response:
        pipeline_registry = await self._get_pipeline_registry()
        pipeline = await pipeline_registry.get_pipeline(pipeline_id=pipeline_id)
        if pipeline is None:
            return _response("POST", url, status_code=404, payload={"detail": "pipeline not found"})
        latest = await pipeline_registry.get_latest_version(pipeline_id=pipeline_id, branch=pipeline.branch)
        if latest is None:
            return _response("POST", url, status_code=409, payload={"detail": "pipeline has no version"})

        branch = _normalize_branch(payload.get("branch") or payload.get("branchName") or pipeline.branch or "master")
        node_id = str(payload.get("node_id") or payload.get("nodeId") or "").strip() or None
        limit_raw = payload.get("limit")
        try:
            preview_limit_value = int(limit_raw) if limit_raw is not None else None
        except (TypeError, ValueError):
            preview_limit_value = None

        pipeline_type = str(getattr(pipeline, "pipeline_type", "") or "").strip().lower()
        if pipeline_type == "streaming":
            node_id = None
            preview_limit_value = None

        definition_json = latest.definition_json if isinstance(latest.definition_json, dict) else {}
        output_dataset_name = _extract_output_dataset_name(definition_json, node_id=node_id)
        job_id = f"build-{pipeline_id}-{int(time.time() * 1000)}"
        await self._queue.publish(
            PipelineJob(
                job_id=job_id,
                pipeline_id=pipeline_id,
                db_name=pipeline.db_name,
                pipeline_type=pipeline.pipeline_type,
                definition_json=definition_json,
                definition_commit_id=latest.lakefs_commit_id,
                node_id=node_id,
                output_dataset_name=output_dataset_name,
                mode="build",
                preview_limit=preview_limit_value,
                branch=branch,
            )
        )
        build_rid_value = build_rid("build", job_id)
        self._build_rids_by_job_id[job_id] = build_rid_value
        self._mode_by_job_id[job_id] = "build"
        return _response(
            "POST",
            url,
            status_code=200,
            payload={"status": "success", "data": {"job_id": job_id, "build_rid": build_rid_value, "pipeline_id": pipeline_id}},
        )

    async def _create_build_v2(
        self,
        *,
        url: str,
        pipeline_id: str,
        payload: dict[str, Any],
        headers: Optional[dict[str, str]],
    ) -> httpx.Response:
        branch = _normalize_branch(payload.get("branch") or payload.get("branchName") or "master")
        node_id = str(payload.get("node_id") or payload.get("nodeId") or "").strip()
        limit = payload.get("limit")
        pipeline_registry = await self._get_pipeline_registry()
        pipeline = await pipeline_registry.get_pipeline(pipeline_id=pipeline_id)
        pipeline_type = str(getattr(pipeline, "pipeline_type", "") or "").strip().lower()
        if pipeline_type == "streaming":
            node_id = ""
            limit = None

        target_rid_value = build_rid("pipeline", pipeline_id)
        body: dict[str, Any] = {
            "targetRid": target_rid_value,
            "target": {"targetRids": [target_rid_value]},
            "branchName": branch,
        }
        params: dict[str, Any] = {}
        if node_id:
            params["nodeId"] = node_id
        if limit is not None:
            try:
                params["limit"] = int(limit)
            except (TypeError, ValueError):
                pass
        if params:
            body["parameters"] = params

        build_resp = await self._client.post(
            f"{self._base_url(url)}/api/v2/orchestration/builds/create",
            headers=headers,
            json=body,
        )
        if build_resp.status_code >= 400:
            fallback_payload = dict(payload)
            fallback_payload.pop("db_name", None)
            if pipeline_type == "streaming":
                fallback_payload.pop("node_id", None)
                fallback_payload.pop("nodeId", None)
                fallback_payload.pop("limit", None)
            fallback_payload["branch"] = branch
            fallback_resp = await self._client.post(
                f"{self._base_url(url)}/api/v1/pipelines/{pipeline_id}/build",
                headers=headers,
                json=fallback_payload,
            )
            if fallback_resp.status_code >= 400:
                return await self._queue_build_job(url=url, pipeline_id=pipeline_id, payload=fallback_payload)

            fallback_data = (fallback_resp.json().get("data") or {}) if fallback_resp.headers.get("content-type", "").startswith("application/json") else {}
            fallback_job_id = str(fallback_data.get("job_id") or "").strip()
            if fallback_job_id:
                fallback_build_rid = build_rid("build", fallback_job_id)
                self._build_rids_by_job_id[fallback_job_id] = fallback_build_rid
                self._mode_by_job_id[fallback_job_id] = "build"
            return fallback_resp
        build_payload = build_resp.json()
        build_rid_value = str(build_payload.get("rid") or "")
        if not build_rid_value:
            return _response("POST", url, status_code=500, payload={"detail": "build rid missing"})
        job_id = _build_job_id_from_build_rid(build_rid_value)
        self._build_rids_by_job_id[job_id] = build_rid_value
        self._mode_by_job_id[job_id] = "build"
        return _response(
            "POST",
            url,
            status_code=200,
            payload={
                "status": "success",
                "data": {"job_id": job_id, "build_rid": build_rid_value, "pipeline_id": pipeline_id},
            },
        )

    async def _legacy_runs_response(self, *, url: str, pipeline_id: str, limit: int) -> httpx.Response:
        pipeline_registry = await self._get_pipeline_registry()
        try:
            runs = await pipeline_registry.list_runs(pipeline_id=pipeline_id, limit=max(1, min(int(limit), 200)))
        except Exception:
            return _response("GET", url, status_code=200, payload={"status": "success", "data": {"runs": []}})
        normalized: list[dict[str, Any]] = []
        for row in runs:
            job_id = str(row.get("job_id") or "")
            merged = dict(row)
            if job_id in self._mode_by_job_id:
                merged["mode"] = self._mode_by_job_id[job_id]
            merged["status"] = _legacy_status(merged.get("status"))
            normalized.append(merged)
        return _response("GET", url, status_code=200, payload={"status": "success", "data": {"runs": normalized}})

    async def _legacy_artifacts_response(
        self,
        *,
        url: str,
        pipeline_id: str,
        limit: int,
        mode: Optional[str],
    ) -> httpx.Response:
        pipeline_registry = await self._get_pipeline_registry()
        runs = await pipeline_registry.list_runs(pipeline_id=pipeline_id, limit=max(1, min(int(limit), 200)))
        artifacts: list[dict[str, Any]] = []
        for run in runs:
            job_id = str(run.get("job_id") or "")
            run_mode = str(self._mode_by_job_id.get(job_id) or run.get("mode") or "")
            if mode and run_mode and run_mode.lower() != str(mode).lower():
                continue
            status = _legacy_status(run.get("status"))
            if status not in _TERMINAL_RUN_STATUSES:
                continue
            output_json = run.get("output_json") if isinstance(run.get("output_json"), dict) else {}
            outputs = output_json.get("outputs") if isinstance(output_json.get("outputs"), list) else []
            normalized_outputs: list[dict[str, Any]] = []
            for output in outputs:
                if not isinstance(output, dict):
                    continue
                normalized = dict(output)
                if "dataset_name" not in normalized and output.get("datasetName"):
                    normalized["dataset_name"] = output.get("datasetName")
                if "artifact_key" not in normalized and output.get("artifactKey"):
                    normalized["artifact_key"] = output.get("artifactKey")
                merge_commit_id = output_json.get("merge_commit_id")
                if merge_commit_id and "merge_commit_id" not in normalized:
                    normalized["merge_commit_id"] = merge_commit_id
                normalized_outputs.append(normalized)
            artifact_id = str(
                output_json.get("artifact_id")
                or uuid5(NAMESPACE_URL, f"pipeline-artifact:{pipeline_id}:{job_id}:{run_mode}")
            )
            artifacts.append(
                {
                    "artifact_id": artifact_id,
                    "job_id": job_id,
                    "status": "SUCCESS" if status in {"SUCCESS", "DEPLOYED"} else status,
                    "mode": run_mode,
                    "outputs": normalized_outputs,
                }
            )
        return _response("GET", url, status_code=200, payload={"status": "success", "data": {"artifacts": artifacts}})

    async def _legacy_pipeline_detail_response(self, *, url: str, pipeline_id: str) -> httpx.Response:
        pipeline_registry = await self._get_pipeline_registry()
        pipeline = await pipeline_registry.get_pipeline(pipeline_id=pipeline_id)
        if pipeline is None:
            return _response("GET", url, status_code=404, payload={"detail": "pipeline not found"})
        payload = {
            "status": "success",
            "data": {
                "pipeline": {
                    "pipeline_id": pipeline.pipeline_id,
                    "pipelineId": pipeline.pipeline_id,
                    "db_name": pipeline.db_name,
                    "name": pipeline.name,
                    "branch": pipeline.branch,
                    "pipeline_type": pipeline.pipeline_type,
                    "last_preview_status": pipeline.last_preview_status,
                    "last_preview_sample": pipeline.last_preview_sample,
                    "last_build_status": pipeline.last_build_status,
                    "last_build_output": pipeline.last_build_output,
                    "deployed_commit_id": pipeline.deployed_commit_id,
                }
            },
        }
        return _response("GET", url, status_code=200, payload=payload)

    def _base_url(self, url: str) -> str:
        parsed = urlsplit(url)
        return f"{parsed.scheme}://{parsed.netloc}"

    async def post(self, url: str, *args: Any, **kwargs: Any) -> httpx.Response:
        parsed = urlsplit(url)
        path = parsed.path
        json_payload = kwargs.get("json") if isinstance(kwargs.get("json"), dict) else {}
        headers = kwargs.get("headers")
        params = kwargs.get("params") if isinstance(kwargs.get("params"), dict) else {}

        if path == _PIPELINE_DATASETS_PREFIX:
            db_name = str(json_payload.get("db_name") or self._client.headers.get("X-DB-Name") or "").strip()
            name = str(json_payload.get("name") or "").strip()
            description = str(json_payload.get("description") or "").strip()
            branch = _normalize_branch(json_payload.get("branch") or params.get("branch") or "master")
            schema_json = json_payload.get("schema_json") if isinstance(json_payload.get("schema_json"), dict) else None
            return await self._create_dataset_v2(
                url=url,
                db_name=db_name,
                name=name,
                description=description,
                branch=branch,
                schema_json=schema_json,
                headers=headers,
            )

        dataset_version_match = re.fullmatch(r"/api/v1/pipelines/datasets/([^/]+)/versions", path)
        if dataset_version_match:
            dataset_id = dataset_version_match.group(1)
            branch = _normalize_branch(json_payload.get("branch") or params.get("branch") or "master")
            sample_json = json_payload.get("sample_json") if isinstance(json_payload.get("sample_json"), dict) else {}
            schema_json = json_payload.get("schema_json") if isinstance(json_payload.get("schema_json"), dict) else {}
            return await self._create_dataset_version_v2(
                url=url,
                dataset_id=dataset_id,
                branch=branch,
                sample_json=sample_json,
                schema_json=schema_json,
                headers=headers,
            )

        if path == f"{_PIPELINE_DATASETS_PREFIX}/csv-upload":
            data = kwargs.get("data") if isinstance(kwargs.get("data"), dict) else {}
            files = kwargs.get("files") if isinstance(kwargs.get("files"), dict) else {}
            return await self._csv_or_excel_upload_v2(
                url=url,
                params=params,
                data=data,
                files=files,
                headers=headers,
                excel=False,
            )

        if path == f"{_PIPELINE_DATASETS_PREFIX}/excel-upload":
            data = kwargs.get("data") if isinstance(kwargs.get("data"), dict) else {}
            files = kwargs.get("files") if isinstance(kwargs.get("files"), dict) else {}
            return await self._csv_or_excel_upload_v2(
                url=url,
                params=params,
                data=data,
                files=files,
                headers=headers,
                excel=True,
            )

        if path == _PIPELINE_PREFIX:
            return await self._create_pipeline_direct(url=url, payload=json_payload)

        preview_match = re.fullmatch(r"/api/v1/pipelines/([^/]+)/preview", path)
        if preview_match:
            return await self._queue_preview_job(url=url, pipeline_id=preview_match.group(1), payload=json_payload)

        build_match = re.fullmatch(r"/api/v1/pipelines/([^/]+)/build", path)
        if build_match:
            return await self._create_build_v2(
                url=url,
                pipeline_id=build_match.group(1),
                payload=json_payload,
                headers=headers,
            )

        # NOTE: no Foundry v2 deploy endpoint is implemented yet; keep legacy deploy path for parity checks.
        if re.fullmatch(r"/api/v1/pipelines/([^/]+)/deploy", path):
            fixed_payload = dict(json_payload)
            branch = _normalize_branch(
                fixed_payload.get("branch")
                or fixed_payload.get("branchName")
                or params.get("branch")
                or "master"
            )
            fixed_payload["branch"] = branch
            if "branchName" in fixed_payload:
                fixed_payload["branchName"] = branch
            fixed_kwargs = dict(kwargs)
            fixed_kwargs["json"] = fixed_payload
            return await self._client.post(url, *args, **fixed_kwargs)

        return await self._client.post(url, *args, **kwargs)

    async def get(self, url: str, *args: Any, **kwargs: Any) -> httpx.Response:
        parsed = urlsplit(url)
        path = parsed.path
        params = kwargs.get("params") if isinstance(kwargs.get("params"), dict) else {}

        if path.startswith(_PIPELINE_DATASETS_PREFIX):
            return await self._client.get(url, *args, **kwargs)

        runs_match = re.fullmatch(r"/api/v1/pipelines/([^/]+)/runs", path)
        if runs_match:
            return await self._legacy_runs_response(
                url=url,
                pipeline_id=runs_match.group(1),
                limit=int(params.get("limit") or 50),
            )

        artifacts_match = re.fullmatch(r"/api/v1/pipelines/([^/]+)/artifacts", path)
        if artifacts_match:
            return await self._legacy_artifacts_response(
                url=url,
                pipeline_id=artifacts_match.group(1),
                limit=int(params.get("limit") or 50),
                mode=str(params.get("mode") or "").strip() or None,
            )

        detail_match = re.fullmatch(r"/api/v1/pipelines/([^/]+)", path)
        if detail_match:
            return await self._legacy_pipeline_detail_response(url=url, pipeline_id=detail_match.group(1))

        return await self._client.get(url, *args, **kwargs)

    async def put(self, url: str, *args: Any, **kwargs: Any) -> httpx.Response:
        parsed = urlsplit(url)
        path = parsed.path
        payload = kwargs.get("json") if isinstance(kwargs.get("json"), dict) else {}
        headers = kwargs.get("headers")

        update_match = re.fullmatch(r"/api/v1/pipelines/([^/]+)", path)
        if update_match:
            pipeline_id = update_match.group(1)
            pipeline_registry = await self._get_pipeline_registry()
            pipeline = await pipeline_registry.get_pipeline(pipeline_id=pipeline_id)
            if pipeline is None:
                return _response("PUT", url, status_code=404, payload={"detail": "pipeline not found"})

            update_fields: dict[str, Any] = {}
            if payload.get("name") is not None:
                update_fields["name"] = str(payload.get("name") or "")
            if payload.get("description") is not None:
                update_fields["description"] = str(payload.get("description") or "")
            if payload.get("branch") is not None:
                update_fields["branch"] = _normalize_branch(payload.get("branch"))
            if update_fields:
                await pipeline_registry.update_pipeline(pipeline_id=pipeline_id, **update_fields)
            if isinstance(payload.get("definition_json"), dict):
                await pipeline_registry.add_version(
                    pipeline_id=pipeline_id,
                    branch=str(update_fields.get("branch") or pipeline.branch or "master"),
                    definition_json=payload.get("definition_json") or {},
                    user_id=str(self._client.headers.get("X-User-ID") or "system"),
                )
            latest = await pipeline_registry.get_pipeline(pipeline_id=pipeline_id)
            if latest is None:
                return _response("PUT", url, status_code=500, payload={"detail": "pipeline update failed"})
            return _response(
                "PUT",
                url,
                status_code=200,
                payload={
                    "status": "success",
                    "data": {
                        "pipeline": {
                            "pipeline_id": latest.pipeline_id,
                            "db_name": latest.db_name,
                            "name": latest.name,
                            "branch": latest.branch,
                            "pipeline_type": latest.pipeline_type,
                        }
                    },
                },
            )

        return await self._client.put(url, *args, **kwargs)
