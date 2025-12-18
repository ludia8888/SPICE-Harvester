import importlib.util
from io import BytesIO

import pytest

from shared.models.structure_analysis import MergeRange
from shared.services.sheet_grid_parser import SheetGridParseOptions, SheetGridParser


class TestSheetGridParser:
    def test_google_values_normalize_and_trim_trailing(self):
        values = [
            ["A", "B"],
            ["1", "2"],
            ["", ""],
            ["3", "4"],
            ["", ""],  # trailing empty row -> should be trimmed
        ]

        out = SheetGridParser.from_google_sheets_values(values)
        assert out.source == "google_sheets"
        assert out.metadata["rows"] == 4
        assert out.metadata["cols"] == 2
        # Internal empty row should be preserved (row index 2)
        assert out.grid[2] == ["", ""]

    def test_google_merges_from_metadata(self):
        meta = {
            "sheets": [
                {
                    "properties": {"title": "Sheet1", "sheetId": 0},
                    "merges": [
                        {
                            "startRowIndex": 0,
                            "endRowIndex": 2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 3,
                        }
                    ],
                }
            ]
        }

        merges = SheetGridParser.merged_cells_from_google_metadata(meta, worksheet_name="Sheet1")
        assert merges == [MergeRange(top=0, left=1, bottom=1, right=2)]

    def test_google_merges_are_clipped_to_grid(self):
        values = [["A"], ["B"]]
        merges = [MergeRange(top=0, left=0, bottom=10, right=10)]

        out = SheetGridParser.from_google_sheets_values(values, merged_cells=merges)
        assert out.metadata["rows"] == 2
        assert out.metadata["cols"] == 1
        # Clipped to the actual grid size (A1:A2)
        assert out.merged_cells == [MergeRange(top=0, left=0, bottom=1, right=0)]

    def test_excel_parser_optional_dependency(self):
        if importlib.util.find_spec("openpyxl") is not None:
            pytest.skip("openpyxl installed; this environment can run full Excel parser tests")

        with pytest.raises(RuntimeError) as exc:
            SheetGridParser.from_excel_bytes(b"not-an-xlsx")
        assert "openpyxl" in str(exc.value).lower()

    @pytest.mark.filterwarnings(
        "ignore:datetime\\.datetime\\.utcnow\\(\\) is deprecated.*:DeprecationWarning:openpyxl\\..*"
    )
    def test_excel_parser_extracts_merges_and_currency_format(self):
        if importlib.util.find_spec("openpyxl") is None:
            pytest.skip("openpyxl not installed")

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A1"] = "상품"
        ws["B1"] = "금액"
        ws["A2"] = "셔츠"
        ws["B2"] = 150
        ws["B2"].number_format = "¥#,##0"

        ws["A3"] = "카테고리"
        ws["A4"] = ""
        ws.merge_cells("A3:A4")

        bio = BytesIO()
        wb.save(bio)

        out = SheetGridParser.from_excel_bytes(
            bio.getvalue(),
            options=SheetGridParseOptions(trim_trailing_empty=True, max_rows=10, max_cols=10),
        )

        assert out.source == "excel"
        assert out.sheet_name == "Sheet1"
        assert out.grid[1][1].startswith("¥")
        assert any(m.top == 2 and m.left == 0 and m.bottom == 3 and m.right == 0 for m in out.merged_cells)
