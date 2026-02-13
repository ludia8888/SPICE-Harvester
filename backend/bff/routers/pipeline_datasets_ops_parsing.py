"""Pipeline dataset parsing helpers.

Small, stable helpers extracted from `bff.routers.pipeline_datasets_ops`.
"""

from __future__ import annotations

import csv
import hashlib
import io
from typing import Any, Dict, Optional

from fastapi import status
from shared.errors.error_types import ErrorCode, classified_http_exception


def _default_dataset_name(filename: str) -> str:
    base = (filename or "").strip()
    if not base:
        return "excel_dataset"
    if "." in base:
        base = ".".join(base.split(".")[:-1]) or base
    return base or "excel_dataset"


def _convert_xls_to_xlsx_bytes(xls_bytes: bytes) -> bytes:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:
        raise classified_http_exception(
            status.HTTP_501_NOT_IMPLEMENTED,
            "XLS support requires the 'xlrd' package.",
            code=ErrorCode.FEATURE_NOT_IMPLEMENTED,
        ) from exc

    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError as exc:
        raise classified_http_exception(
            status.HTTP_501_NOT_IMPLEMENTED,
            "XLS support requires the 'openpyxl' package.",
            code=ErrorCode.FEATURE_NOT_IMPLEMENTED,
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=xls_bytes)
    except Exception as exc:
        raise classified_http_exception(status.HTTP_400_BAD_REQUEST, f"Failed to read .xls file: {exc}", code=ErrorCode.REQUEST_VALIDATION_FAILED) from exc

    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove(workbook.worksheets[0])

    for sheet in book.sheets():
        title = (sheet.name or "Sheet1").strip() or "Sheet1"
        worksheet = workbook.create_sheet(title=title[:31])
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                worksheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _normalize_table_bbox(
    *,
    table_top: Optional[int],
    table_left: Optional[int],
    table_bottom: Optional[int],
    table_right: Optional[int],
) -> Optional[Dict[str, int]]:
    parts = [table_top, table_left, table_bottom, table_right]
    if all(part is None for part in parts):
        return None
    if any(part is None for part in parts):
        raise classified_http_exception(
            status.HTTP_400_BAD_REQUEST,
            "table_top/table_left/table_bottom/table_right must be provided together",
            code=ErrorCode.REQUEST_VALIDATION_FAILED,
        )
    return {
        "top": int(table_top),
        "left": int(table_left),
        "bottom": int(table_bottom),
        "right": int(table_right),
    }


def _detect_csv_delimiter(sample: str) -> str:
    candidates = [",", "\t", ";", "|"]
    line = next((line for line in sample.splitlines() if line.strip()), "")
    if not line:
        return ","
    scores = {delimiter: line.count(delimiter) for delimiter in candidates}
    best = max(scores, key=scores.get)
    return best if scores.get(best, 0) > 0 else ","


def _parse_csv_rows(
    reader: Any,
    *,
    has_header: bool,
    preview_limit: int,
) -> tuple[list[str], list[list[str]], int]:
    columns: list[str] = []
    preview_rows: list[list[str]] = []
    total_rows = 0

    for row in reader:
        if not row or all(str(cell).strip() == "" for cell in row):
            continue
        if not columns:
            if has_header:
                columns = [str(cell).strip() or f"column_{index + 1}" for index, cell in enumerate(row)]
                continue
            columns = [f"column_{index + 1}" for index in range(len(row))]
        normalized = [str(cell).strip() for cell in row]
        if len(normalized) < len(columns):
            normalized.extend([""] * (len(columns) - len(normalized)))
        elif len(normalized) > len(columns):
            normalized = normalized[: len(columns)]
        total_rows += 1
        if len(preview_rows) < preview_limit:
            preview_rows.append(normalized)

    return columns, preview_rows, total_rows


def _parse_csv_file(
    file_obj: Any,
    *,
    delimiter: str,
    has_header: bool,
    preview_limit: int = 200,
) -> tuple[list[str], list[list[str]], int, str]:
    if not hasattr(file_obj, "read"):
        raise ValueError("file object does not support read()")
    if not hasattr(file_obj, "seek"):
        raise ValueError("file object does not support seek()")

    class _HashingReader(io.RawIOBase):
        def __init__(self, raw: Any, hasher: "hashlib._Hash") -> None:
            self._raw = raw
            self._hasher = hasher

        def read(self, size: int = -1) -> bytes:
            chunk = self._raw.read(size)
            if chunk:
                self._hasher.update(chunk)
            return chunk

        def readinto(self, b: bytearray | memoryview) -> int:
            chunk = self.read(len(b))
            if not chunk:
                return 0
            b[: len(chunk)] = chunk
            return len(chunk)

        def readable(self) -> bool:
            return True

        def seekable(self) -> bool:
            return hasattr(self._raw, "seek")

        def seek(self, offset: int, whence: int = io.SEEK_SET) -> int:
            return self._raw.seek(offset, whence)

        def tell(self) -> int:
            return self._raw.tell()

    file_obj.seek(0)
    hasher = hashlib.sha256()
    hashing_reader = _HashingReader(file_obj, hasher)
    buffered = io.BufferedReader(hashing_reader)
    text_stream = io.TextIOWrapper(buffered, encoding="utf-8", errors="replace", newline="")

    try:
        reader = csv.reader(text_stream, delimiter=delimiter)
        columns, preview_rows, total_rows = _parse_csv_rows(
            reader,
            has_header=has_header,
            preview_limit=preview_limit,
        )
    finally:
        try:
            text_stream.detach()
        except Exception:
            pass
        try:
            file_obj.seek(0)
        except Exception:
            pass

    return columns, preview_rows, total_rows, hasher.hexdigest()


def _parse_csv_content(
    content: str,
    *,
    delimiter: str,
    has_header: bool,
    preview_limit: int = 200,
) -> tuple[list[str], list[list[str]], int]:
    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    return _parse_csv_rows(
        reader,
        has_header=has_header,
        preview_limit=preview_limit,
    )
